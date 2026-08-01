#!/usr/bin/env python3
# ===========================================================================
# AS-RUN PROVENANCE COPY — frozen 2026-07-31 (decision-surface audit gap 3:
# the seven-domain external-read runners were previously uncommitted).
# Source:        /mnt/v/output/zensim/rousselot-chroma-2026-07-29/analyze_rousselot.py
# sha256(source): e7028e5b1ac041fc66b7a763d539e8645f01f203038b977353d053622c819749
# build_commit:  73734d8820b46c825aea26f8e4511d50e6a92dc7
# Protocol doc:  benchmarks/rousselot_chroma_validation_2026-07-29.md
# Everything below the marker line is BYTE-IDENTICAL to the source file
# (verify: strip through the marker, sha256 the rest). Do NOT extend this
# file — it is an archival record of the exact as-run analysis (it may call
# scipy directly; it predates the stats-rule batch migration and is kept
# verbatim). New analyses go through scripts/external_reads/run_external_reads.py.
# Data paths inside are as-run: /mnt/v artifact paths persist; ~/tmp
# FEATS_DIR-style inputs were session scratch — the pooled/persisted tables
# in the artifact dir are the stored equivalents (see ../README.md).
# ===== byte-identical source below this line ================================
"""Rousselot HDdtb/4Kdtb chroma blind-spot validation — analysis per
PROTOCOL.md (pre-registered; registered looks L1/L2/L3 + Q3 diagnostic +
comparators + K=100 robustness). Machinery = the SI-HDR study's
analyze_sihdr.py nested-CV/bootstrap code adapted to LOSO(8).

Inputs (this dir): {hddtb,k4dtb}_feats_944_k{179,100}.csv, pairs_manifest.json
Outputs: results.json + stdout log. Seed 20260729 everywhere.
"""
import csv
import json
import os
import re
from collections import defaultdict

import numpy as np
from scipy.stats import spearmanr, kendalltau
from sklearn.linear_model import Ridge
from sklearn.model_selection import GroupKFold

BASE = "/mnt/v/output/zensim/rousselot-chroma-2026-07-29"
SEED = 20260729
LAMBDAS = [1e-2, 1e-1, 1.0, 10.0, 100.0]
NBOOT = 10000
OUT = {}

def sr(a, b):
    return float(spearmanr(a, b).statistic)

# ---------------- registered feature-subset index lists ----------------
def subset_indices():
    yonly, chroma = [], []
    for s in range(4):
        for ch in range(3):
            b = s * 39 + ch * 13
            (yonly if ch == 1 else chroma).extend(range(b, b + 13))
    for s in range(4):
        for ch in range(3):
            b = 372 + s * 87 + ch * 29
            (yonly if ch == 1 else chroma).extend(range(b, b + 29))
    for s in range(4):
        for ch in range(3):
            b = 720 + s * 51 + ch * 17
            if ch == 1:
                chroma.append(b)            # XMASK / K_XCH lane
                yonly.extend(range(b + 1, b + 17))
            else:
                chroma.extend(range(b, b + 17))
    yonly.extend(range(924, 944))
    return np.array(sorted(yonly)), np.array(sorted(chroma))

Y_IDX, C_IDX = subset_indices()
assert len(Y_IDX) == 252 and len(C_IDX) == 476
assert not (set(Y_IDX) & set(C_IDX))
assert set(range(156, 372)).isdisjoint(set(Y_IDX) | set(C_IDX))
OUT["subsets"] = {"yonly_n": 252, "chroma_n": 476, "dead_n": 216,
                  "yonly_idx_sha": hash(tuple(Y_IDX)) & 0xffffffff,
                  "chroma_idx_sha": hash(tuple(C_IDX)) & 0xffffffff}

# ---------------- lane names for Q3 ----------------
V1N = ["ssim_mean","ssim_4th","ssim_2nd","edge_art_mean","edge_art_4th",
       "edge_art_2nd","edge_det_mean","edge_det_4th","edge_det_2nd","mse",
       "var_loss","tex_loss","contrast_inc"]
V2N = ["SSIM_MEAN","SSIM_DEV2","SSIM_DEV4","ART","DET","MSE","HF_GAIN",
       "HF_LOSS","HF_MAG_LOSS","SSIM_SOFT_PEAK","ART_SOFT_PEAK","DET_SOFT_PEAK",
       "MASKED_SSIM","MASKED_ART","MASKED_DET","MASKED_MSE","IW_SSIM","IW_ART",
       "IW_DET","IW_MSE","PJND_TRANSDUCER","PJND_FRAGILITY","GMS",
       "PJND_TRANSDUCER_LOW_K","PJND_TRANSDUCER_HIGH_K","BLOCKINESS","RINGING",
       "BANDING","EDGE_WIDTH_CHANGE"]
APPN = ["XMASK_TRANSDUCER","LUM_TRANSDUCER","LUM_DARK_ERR","LUM_MID_ERR",
        "LUM_BRIGHT_ERR","MSCN_DIFF_MEAN","MSCN_DIFF_L2","CONTRAST_GAIN",
        "CONTRAST_LOSS","TEXTURE_DISSIM","GMS_DEV2","ART_DEV2","DET_DEV2",
        "GLOBAL_DMEAN","GLOBAL_CGAIN","GLOBAL_CLOSS","GRAD_SRC_MEAN"]
APP2N = ["BANDVIS_GAIN","BANDVIS_LOSS","LUMA_MEAN_REF","HL_BIN1","HL_BIN2"]
CH = ["X","Y","B"]

def lane_name(i):
    if i < 156:
        s, r = divmod(i, 39); ch, k = divmod(r, 13)
        return f"v1.s{s}.{CH[ch]}.{V1N[k]}", ("K_XCH" if False else f"{CH[ch]}-v1")
    if i < 372:
        return f"dead.f{i}", "dead"
    if i < 720:
        j = i - 372; s, r = divmod(j, 87); ch, k = divmod(r, 29)
        return f"v2.s{s}.{CH[ch]}.{V2N[k]}", f"{CH[ch]}-v2"
    if i < 924:
        j = i - 720; s, r = divmod(j, 51); ch, k = divmod(r, 17)
        grp = "K_XCH" if (ch == 1 and k == 0) else f"{CH[ch]}-append"
        return f"app.s{s}.{CH[ch]}.{APPN[k]}", grp
    j = i - 924; s, k = divmod(j, 5)
    return f"app2.s{s}.Y.{APP2N[k]}", "append2-Y"

# ---------------- load ----------------
def load_feats(path):
    with open(path, newline="") as f:
        rdr = csv.reader(f); header = next(rdr); rows = list(rdr)
    off = header.index("f0")
    assert header[off:] == [f"f{k}" for k in range(944)]
    ids = [r[0] for r in rows]
    X = np.array([[float(v) for v in r[off:]] for r in rows])
    s228 = np.array([float(r[header.index("score228")]) for r in rows])
    clamp = np.array([[float(r[header.index("clampfrac_ref")]),
                       float(r[header.index("clampfrac_dist")])] for r in rows])
    return ids, X, s228, clamp

man = json.load(open(f"{BASE}/pairs_manifest.json"))  # built from the xlsx + archives; sha-matched copy of the build-time scratch manifest
meta = {}
for r in man:
    rid = r["dist"].split("/")[-1].replace(".hdr", "")
    meta[rid] = r

DS = {}
for ds, fn in [("hddtb", "hddtb_feats_944_k179.csv"), ("4kdtb", "k4dtb_feats_944_k179.csv")]:
    ids, X, s228, clamp = load_feats(f"{BASE}/{fn}")
    mos = np.array([meta[i]["mos"] for i in ids])
    scene = np.array([meta[i]["scene"] for i in ids])
    fam = np.array([meta[i]["family"] for i in ids])
    DS[ds] = dict(ids=ids, X=X, s228=s228, clamp=clamp, mos=mos, scene=scene, fam=fam)
    print(f"{ds}: {len(ids)} rows, scenes={len(set(scene))}, fams={sorted(set(fam))}, "
          f"max clampfrac ref/dist = {clamp[:,0].max():.4f}/{clamp[:,1].max():.4f}")
OUT["coverage"] = {ds: {"rows": len(d["ids"]), "scenes": len(set(d["scene"])),
                        "max_clampfrac_ref": float(d["clamp"][:,0].max()),
                        "max_clampfrac_dist": float(d["clamp"][:,1].max())}
                   for ds, d in DS.items()}

# chroma-pure mask (HDdtb)
cp_mask = np.isin(DS["hddtb"]["fam"], ["cnoise", "gamut"])
assert cp_mask.sum() == 40
OUT["coverage"]["hddtb"]["chroma_pure_rows"] = int(cp_mask.sum())
m = DS["hddtb"]["mos"][cp_mask]
print(f"HDdtb chroma-pure MOS: n=40 range {m.min():.1f}..{m.max():.1f} std {m.std():.1f}")
OUT["coverage"]["hddtb"]["chroma_pure_mos_std"] = float(m.std())

# ---------------- LOSO nested CV OOF (template machinery, LOSO(8)) ------
def nested_cv_oof(X, y, scene_groups, col_idx=None):
    Xc = X if col_idx is None else X[:, col_idx]
    oof = np.full(len(y), np.nan)
    lam_picks = []
    outer = GroupKFold(n_splits=len(set(scene_groups)))
    for tr, te in outer.split(Xc, y, scene_groups):
        keep = Xc[tr].std(axis=0) > 1e-12
        Xk = Xc[:, keep]
        mu, sd = Xk[tr].mean(axis=0), Xk[tr].std(axis=0)
        Ztr, Zte = (Xk[tr] - mu) / sd, (Xk[te] - mu) / sd
        inner = GroupKFold(n_splits=4)
        cv = {}
        for lam in LAMBDAS:
            scores = []
            for itr, ite in inner.split(Ztr, y[tr], scene_groups[tr]):
                mdl = Ridge(alpha=lam)
                mdl.fit(Ztr[itr], y[tr][itr])
                scores.append(sr(mdl.predict(Ztr[ite]), y[tr][ite]))
            cv[lam] = float(np.mean(scores))
        lam = max(cv, key=cv.get)
        lam_picks.append(lam)
        mdl = Ridge(alpha=lam)
        mdl.fit(Ztr, y[tr])
        oof[te] = mdl.predict(Zte)
    assert not np.isnan(oof).any()
    return oof, lam_picks

def paired_row_bootstrap(pa, pb, y, n=NBOOT, seed=SEED):
    """Δ = sr(pa)−sr(pb) row bootstrap; returns mean, ci95, p(Δ<=0), p(Δ>=0)."""
    rng = np.random.default_rng(seed)
    deltas = []
    idx_all = np.arange(len(y))
    bad = 0
    while len(deltas) < n:
        idx = rng.choice(idx_all, size=len(y), replace=True)
        if len(set(y[idx])) < 2 or len(set(pa[idx])) < 2 or len(set(pb[idx])) < 2:
            bad += 1
            if bad > n:  # give up guard
                break
            continue
        deltas.append(sr(pa[idx], y[idx]) - sr(pb[idx], y[idx]))
    d = np.array(deltas)
    return dict(mean=float(d.mean()), ci95=[float(np.percentile(d, 2.5)),
                float(np.percentile(d, 97.5))],
                p_le0=float((d <= 0).mean()), p_ge0=float((d >= 0).mean()),
                degenerate_resamples=bad)

def paired_scene_bootstrap(pa, pb, y, scenes, n=NBOOT, seed=SEED):
    rng = np.random.default_rng(seed)
    uniq = sorted(set(scenes))
    by_scene = {s: np.where(scenes == s)[0] for s in uniq}
    deltas = []
    bad = 0
    while len(deltas) < n:
        pick = rng.choice(uniq, size=len(uniq), replace=True)
        idx = np.concatenate([by_scene[s] for s in pick])
        if len(set(y[idx])) < 2 or len(set(pa[idx])) < 2 or len(set(pb[idx])) < 2:
            bad += 1
            if bad > n:
                break
            continue
        deltas.append(sr(pa[idx], y[idx]) - sr(pb[idx], y[idx]))
    d = np.array(deltas)
    return dict(mean=float(d.mean()), ci95=[float(np.percentile(d, 2.5)),
                float(np.percentile(d, 97.5))],
                p_le0=float((d <= 0).mean()), p_ge0=float((d >= 0).mean()),
                degenerate_resamples=bad)

SUBSETS = {"full944": None, "yonly": Y_IDX, "chroma": C_IDX}

# ---------------- L1+L2 main loop (K=179 verdict extraction) ----------------
OOF = {}
for ds in ["hddtb", "4kdtb"]:
    d = DS[ds]
    OOF[ds] = {}
    for name, idx in SUBSETS.items():
        oof, lams = nested_cv_oof(d["X"], d["mos"], d["scene"], idx)
        OOF[ds][name] = oof
        OUT.setdefault("lambda_picks", {}).setdefault(ds, {})[name] = lams
        print(f"[{ds}] {name}: LOSO OOF SROCC(all {len(oof)}) = {sr(oof, d['mos']):+.4f}  λ={lams}")

l1 = {}
d = DS["hddtb"]
for name in SUBSETS:
    l1[f"hddtb_all_{name}"] = sr(OOF["hddtb"][name], d["mos"])
    l1[f"hddtb_chromapure_{name}"] = sr(OOF["hddtb"][name][cp_mask], d["mos"][cp_mask])
for name in SUBSETS:
    l1[f"4kdtb_all_{name}"] = sr(OOF["4kdtb"][name], DS["4kdtb"]["mos"])

# THE decisive statistic
pa, pb = OOF["hddtb"]["full944"][cp_mask], OOF["hddtb"]["yonly"][cp_mask]
ycp, scp = d["mos"][cp_mask], d["scene"][cp_mask]
l1["delta_full_minus_yonly_chromapure"] = sr(pa, ycp) - sr(pb, ycp)
l1["boot_row"] = paired_row_bootstrap(pa, pb, ycp)
l1["boot_scene"] = paired_scene_bootstrap(pa, pb, ycp, scp)
# attribution
pc = OOF["hddtb"]["chroma"][cp_mask]
l1["boot_row_chroma_vs_yonly"] = paired_row_bootstrap(pc, pb, ycp)
print("\n=== L1 decisive (HDdtb chroma-pure n=40) ===")
for name in SUBSETS:
    print(f"  {name:8s} SROCC = {l1[f'hddtb_chromapure_{name}']:+.4f}")
print(f"  Δ(full−yonly) = {l1['delta_full_minus_yonly_chromapure']:+.4f}  "
      f"row-boot p(Δ≤0)={l1['boot_row']['p_le0']:.4f} ci95={l1['boot_row']['ci95']}  "
      f"scene-boot p(Δ≤0)={l1['boot_scene']['p_le0']:.4f} ci95={l1['boot_scene']['ci95']}")

# 4Kdtb method-contrast: per (scene,Qp) Kendall tau over the 3 treatments
def qp_of(rid):
    m = re.search(r"_Qp(\d+)", rid)
    return int(m.group(1))

d4 = DS["4kdtb"]
blocks = defaultdict(list)
for j, rid in enumerate(d4["ids"]):
    blocks[(d4["scene"][j], qp_of(rid))].append(j)
assert len(blocks) == 32 and all(len(v) == 3 for v in blocks.values())

def block_taus(pred):
    taus = []
    skipped = 0
    for k, idxs in sorted(blocks.items()):
        t = kendalltau(pred[idxs], d4["mos"][idxs]).statistic
        if np.isnan(t):
            skipped += 1
            continue
        taus.append(t)
    return np.array(taus), skipped

def paired_block_bootstrap(pa, pb, n=NBOOT, seed=SEED):
    keys = sorted(blocks.keys())
    ta = {}; tb = {}
    for k in keys:
        idxs = blocks[k]
        ta[k] = kendalltau(pa[idxs], d4["mos"][idxs]).statistic
        tb[k] = kendalltau(pb[idxs], d4["mos"][idxs]).statistic
    keys = [k for k in keys if not (np.isnan(ta[k]) or np.isnan(tb[k]))]
    rng = np.random.default_rng(seed)
    deltas = []
    for _ in range(n):
        pick = rng.choice(len(keys), size=len(keys), replace=True)
        deltas.append(np.mean([ta[keys[i]] for i in pick]) -
                      np.mean([tb[keys[i]] for i in pick]))
    dd = np.array(deltas)
    return dict(mean=float(dd.mean()), ci95=[float(np.percentile(dd, 2.5)),
                float(np.percentile(dd, 97.5))],
                p_le0=float((dd <= 0).mean()), p_ge0=float((dd >= 0).mean()),
                n_blocks=len(keys))

l1["4kdtb_block_tau"] = {}
for name in SUBSETS:
    taus, skipped = block_taus(OOF["4kdtb"][name])
    l1["4kdtb_block_tau"][name] = dict(mean=float(taus.mean()), n=len(taus),
                                       skipped_tied=skipped)
    print(f"  4kdtb mean block-τ ({name}): {taus.mean():+.4f} over {len(taus)} blocks"
          + (f" ({skipped} tied-skip)" if skipped else ""))
l1["4kdtb_block_tau"]["boot_full_vs_yonly"] = paired_block_bootstrap(
    OOF["4kdtb"]["full944"], OOF["4kdtb"]["yonly"])
print(f"  4kdtb Δτ(full−yonly) boot: {l1['4kdtb_block_tau']['boot_full_vs_yonly']}")

# HDdtb matched-(scene,Qp) dist vs distc sign agreement
dh = DS["hddtb"]
def hd_variant(rid):
    if "-distc_Qp" in rid: return "distc", qp_of(rid)
    if "-dist_Qp" in rid: return "dist", qp_of(rid)
    return None, None

pairs = []
by_key = {}
for j, rid in enumerate(dh["ids"]):
    v, qp = hd_variant(rid)
    if v: by_key[(dh["scene"][j], qp, v)] = j
for (sc, qp, v), j in by_key.items():
    if v == "dist" and (sc, qp, "distc") in by_key:
        pairs.append((j, by_key[(sc, qp, "distc")]))
l1["hddtb_matched_pairs_n"] = len(pairs)

def sign_agreement(pred):
    agree = 0; ties = 0
    for a, b in pairs:
        dm = dh["mos"][a] - dh["mos"][b]
        dp = pred[a] - pred[b]
        if dm == 0 or dp == 0: ties += 1; continue
        agree += int(np.sign(dm) == np.sign(dp))
    return agree, ties

for name in SUBSETS:
    agree, ties = sign_agreement(OOF["hddtb"][name])
    l1[f"hddtb_signagree_{name}"] = dict(agree=agree, of=len(pairs) - ties, ties=ties)
    print(f"  hddtb dist-vs-distc sign agreement ({name}): {agree}/{len(pairs)-ties}")
OUT["l1"] = l1

# ---------------- L2 per-family table ----------------
print("\n=== L2 per-family SROCC (OOF) ===")
l2 = {}
for ds in ["hddtb", "4kdtb"]:
    d = DS[ds]
    for fam in sorted(set(d["fam"])):
        mask = d["fam"] == fam
        row = {"n": int(mask.sum())}
        for name in SUBSETS:
            row[name] = sr(OOF[ds][name][mask], d["mos"][mask])
        row["score228"] = sr(d["s228"][mask], d["mos"][mask])
        l2[f"{ds}.{fam}"] = row
        print(f"  {ds}.{fam:14s} n={row['n']:3d} full={row['full944']:+.4f} "
              f"yonly={row['yonly']:+.4f} chroma={row['chroma']:+.4f} "
              f"score228={row['score228']:+.4f}")
OUT["l2"] = l2

# ---------------- comparators ----------------
print("\n=== comparators (zero-fit) ===")
comp = {}
for ds in ["hddtb", "4kdtb"]:
    d = DS[ds]
    comp[f"{ds}_all_score228"] = sr(d["s228"], d["mos"])
comp["hddtb_chromapure_score228"] = sr(DS["hddtb"]["s228"][cp_mask],
                                       DS["hddtb"]["mos"][cp_mask])
taus, skipped = block_taus(d4["s228"])
comp["4kdtb_block_tau_score228"] = dict(mean=float(taus.mean()), n=len(taus))
agree, ties = sign_agreement(dh["s228"])
comp["hddtb_signagree_score228"] = dict(agree=agree, of=len(pairs) - ties)
# display-nits RMSE comparator (cached)
rmse_path = f"{BASE}/rousselot_rmse.csv"
if os.path.exists(rmse_path):
    rmse = {}
    with open(rmse_path) as f:
        for line in f:
            k, v = line.strip().split(",")
            rmse[k] = float(v)
    for ds in ["hddtb", "4kdtb"]:
        d = DS[ds]
        nr = np.array([-rmse[i] for i in d["ids"]])
        comp[f"{ds}_all_negrmse"] = sr(nr, d["mos"])
    nrh = np.array([-rmse[i] for i in DS["hddtb"]["ids"]])
    comp["hddtb_chromapure_negrmse"] = sr(nrh[cp_mask], DS["hddtb"]["mos"][cp_mask])
    taus, _ = block_taus(np.array([-rmse[i] for i in d4["ids"]]))
    comp["4kdtb_block_tau_negrmse"] = float(taus.mean())
for k, v in comp.items():
    print(f"  {k}: {v}")
OUT["comparators"] = comp

# ---------------- L3 expert MOS (secondary, one look) ----------------
print("\n=== L3 expert MOS (HDdtb, secondary/non-verdict) ===")
import openpyxl
wb = openpyxl.load_workbook(
    "/mnt/v/datasets/rousselot-hdr/hddtb-labels/Resultats_Tests_subj.xlsx",
    data_only=True)
we = wb["Expert_MOS"]
grp_norm = {"Gaussian noise": "Gaussian noise", "gamut mismatch": "gamut mismatch",
            "Compression avec chroma Qp": "Compression with chroma Qp algorithm",
            "Compression sans chroma Qp": "Compression without chroma Qp algorithm"}
exp_rows = {}
cur_src = cur_grp = None
for r in range(1, we.max_row + 1):
    src = we.cell(row=r, column=1).value
    grp = we.cell(row=r, column=2).value
    lvl = we.cell(row=r, column=3).value
    mosv = we.cell(row=r, column=13).value
    if src: cur_src = str(src).strip()
    if grp:
        g = str(grp).strip()
        cur_grp = grp_norm.get(g, g)
    if lvl is None or mosv is None: continue
    try:
        mosf = float(mosv)
    except (TypeError, ValueError):
        continue  # repeated in-sheet header rows (PROTOCOL Deviations #1)
    exp_rows[(cur_src, cur_grp, str(lvl).replace(" ", "").replace(",", "."))] = mosf
# naive-side keys from the naive sheet parse (manifest levels/groups)
grp_of_fam = {"cnoise": "Gaussian noise", "gamut": "gamut mismatch",
              "hevc_cqp": "Compression with chroma Qp algorithm",
              "hevc_nocqp": "Compression without chroma Qp algorithm"}
exp_mos = np.full(len(dh["ids"]), np.nan)
for j, rid in enumerate(dh["ids"]):
    r = meta[rid]
    key = (r["scene"], grp_of_fam[r["family"]],
           r["level"].replace(" ", "").replace(",", "."))
    if key in exp_rows:
        exp_mos[j] = exp_rows[key]
matched = int(np.isfinite(exp_mos).sum())
print(f"  expert rows matched: {matched}/96 (keys normalized on (scene,group,level))")
l3 = {"matched": matched}
if matched >= 90:
    ok = np.isfinite(exp_mos)
    cpx = cp_mask & ok
    for name in SUBSETS:
        l3[f"chromapure_{name}"] = sr(OOF["hddtb"][name][cpx], exp_mos[cpx])
    pa2, pb2 = OOF["hddtb"]["full944"][cpx], OOF["hddtb"]["yonly"][cpx]
    l3["delta_full_minus_yonly"] = (sr(pa2, exp_mos[cpx]) - sr(pb2, exp_mos[cpx]))
    l3["boot_row"] = paired_row_bootstrap(pa2, pb2, exp_mos[cpx])
    l3["naive_vs_expert_mos_srocc_chromapure"] = sr(dh["mos"][cpx], exp_mos[cpx])
    l3["naive_vs_expert_mos_srocc_all"] = sr(dh["mos"][ok], exp_mos[ok])
    print(f"  chroma-pure vs EXPERT MOS: " +
          " ".join(f"{n}={l3[f'chromapure_{n}']:+.4f}" for n in SUBSETS))
    print(f"  Δ(full−yonly)={l3['delta_full_minus_yonly']:+.4f} "
          f"p(Δ≤0)={l3['boot_row']['p_le0']:.4f}")
    print(f"  naive-vs-expert MOS SROCC: all={l3['naive_vs_expert_mos_srocc_all']:+.4f} "
          f"chromapure={l3['naive_vs_expert_mos_srocc_chromapure']:+.4f}")
else:
    print("  MATCH FAILED (<90) — L3 dropped per protocol, mismatch documented")
OUT["l3_expert"] = l3

# ---------------- Q3 lane attribution (zero-fit) ----------------
print("\n=== Q3 per-lane |SROCC| (zero-fit, diagnostic) ===")
def lane_scan(X, y, tag, topn=15):
    out = []
    for i in range(944):
        col = X[:, i]
        std = col.std()
        if std <= 1e-12: continue
        s = sr(col, y)
        if np.isnan(s): continue
        nm, grp = lane_name(i)
        out.append((abs(s), s, i, nm, grp, std))
    out.sort(reverse=True)
    print(f"  -- {tag}: top {topn} of {len(out)} live lanes")
    for a, s, i, nm, grp, std in out[:topn]:
        print(f"    f{i:3d} {nm:34s} [{grp:9s}] SROCC={s:+.4f} std={std:.3g}")
    grp_best = {}
    for a, s, i, nm, grp, std in out:
        if grp not in grp_best: grp_best[grp] = (a, s, i, nm)
    return {"top": [dict(f=i, name=nm, group=grp, srocc=s, std=std)
                    for a, s, i, nm, grp, std in out[:topn]],
            "group_best": {g: dict(f=i, name=nm, abs_srocc=a, srocc=s)
                           for g, (a, s, i, nm) in sorted(grp_best.items())},
            "n_live": len(out)}

q3 = {}
q3["hddtb_chromapure"] = lane_scan(DS["hddtb"]["X"][cp_mask], DS["hddtb"]["mos"][cp_mask],
                                   "HDdtb chroma-pure (n=40)")
q3["4kdtb_all"] = lane_scan(DS["4kdtb"]["X"], DS["4kdtb"]["mos"], "4Kdtb all (n=96)")
print("  group-best (HDdtb chroma-pure):")
for g, v in q3["hddtb_chromapure"]["group_best"].items():
    print(f"    {g:10s} best |SROCC|={v['abs_srocc']:.4f} f{v['f']} {v['name']}")
OUT["q3"] = q3

# ---------------- K=100 robustness (registered, non-verdict) ----------------
print("\n=== robustness: K=100 extraction (L1 primary recompute) ===")
rb = {}
for ds, fn in [("hddtb", "hddtb_feats_944_k100.csv"), ("4kdtb", "k4dtb_feats_944_k100.csv")]:
    ids, X, s228, clamp = load_feats(f"{BASE}/{fn}")
    assert ids == DS[ds]["ids"]
    d = DS[ds]
    o = {}
    for name in ["full944", "yonly"]:
        oof, _ = nested_cv_oof(X, d["mos"], d["scene"], SUBSETS[name])
        o[name] = oof
    if ds == "hddtb":
        rb["hddtb_chromapure_full"] = sr(o["full944"][cp_mask], d["mos"][cp_mask])
        rb["hddtb_chromapure_yonly"] = sr(o["yonly"][cp_mask], d["mos"][cp_mask])
        rb["delta"] = rb["hddtb_chromapure_full"] - rb["hddtb_chromapure_yonly"]
        rb["boot_row"] = paired_row_bootstrap(o["full944"][cp_mask], o["yonly"][cp_mask],
                                              d["mos"][cp_mask])
    else:
        ta, _ = block_taus(o["full944"]); tb, _ = block_taus(o["yonly"])
        rb["4kdtb_block_tau_full"] = float(ta.mean())
        rb["4kdtb_block_tau_yonly"] = float(tb.mean())
print(f"  K=100: chromapure full={rb['hddtb_chromapure_full']:+.4f} "
      f"yonly={rb['hddtb_chromapure_yonly']:+.4f} Δ={rb['delta']:+.4f} "
      f"p(Δ≤0)={rb['boot_row']['p_le0']:.4f}")
print(f"  K=100: 4kdtb block-τ full={rb['4kdtb_block_tau_full']:+.4f} "
      f"yonly={rb['4kdtb_block_tau_yonly']:+.4f}")
OUT["robustness_k100"] = rb

json.dump(OUT, open(f"{BASE}/results.json", "w"), indent=1, default=float)
print(f"\nresults.json written")
